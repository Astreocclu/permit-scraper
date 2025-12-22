#!/usr/bin/env python3
# scrapers/ellis_county_excel.py
"""
ELLIS COUNTY PERMIT SCRAPER (Excel)
Source: Ellis County Department of Development
URL: https://www.co.ellis.tx.us/index.aspx?nid=1074

Downloads and parses monthly Excel reports of building permits.
NOTE: Only covers UNINCORPORATED areas (not cities like Waxahachie, Ennis, Midlothian).

Usage:
    python3 scrapers/ellis_county_excel.py                  # Show discovery info
    python3 scrapers/ellis_county_excel.py --url <url>      # Download and parse
    python3 scrapers/ellis_county_excel.py --file data.xlsx # Parse local file
"""

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import requests

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    exit(1)

OUTPUT_DIR = Path(__file__).parent.parent / "data" / "raw"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Known column name variations (all lowercase for matching)
COLUMN_ALIASES = {
    'permit_id': ['permit #', 'permit number', 'permit no', 'permit', 'permit id'],
    'address': ['address', 'location', 'property address', 'site address'],
    'date': ['date', 'issue date', 'issued date', 'permit date'],
    'type': ['type', 'permit type', 'description', 'work type'],
    'value': ['value', 'valuation', 'est value', 'estimated value', 'cost'],
    'owner': ['owner', 'owner name', 'property owner'],
    'contractor': ['contractor', 'contractor name', 'builder'],
}


def detect_column_mapping(headers: List[str]) -> Dict[str, int]:
    """
    Dynamically detect column positions from header row.

    Handles Excel files where columns may be in different positions
    or have slightly different names.
    """
    mapping = {}
    headers_lower = [str(h).lower().strip() if h else '' for h in headers]

    for field, aliases in COLUMN_ALIASES.items():
        for i, header in enumerate(headers_lower):
            if header in aliases:
                mapping[field] = i
                break

    return mapping


def find_header_row(sheet) -> tuple:
    """
    Find the header row in an Excel sheet.

    Ellis County files sometimes have metadata rows before the actual headers.
    Look for a row containing expected column names.
    """
    expected_keywords = ['permit', 'address', 'date']

    for row_idx in range(1, min(10, sheet.max_row + 1)):
        row = [cell.value for cell in sheet[row_idx]]
        row_lower = [str(c).lower() if c else '' for c in row]

        matches = sum(1 for kw in expected_keywords if any(kw in cell for cell in row_lower))
        if matches >= 2:
            return row_idx, row

    row = [cell.value for cell in sheet[1]]
    return 1, row


def parse_excel_permits(file_path: Path) -> List[dict]:
    """
    Parse permits from an Ellis County Excel file.

    Handles dynamic column positions and header row detection.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"ERROR: Failed to load Excel file: {e}")
        return []
    sheet = workbook.active

    header_row_idx, headers = find_header_row(sheet)
    mapping = detect_column_mapping(headers)

    if 'permit_id' not in mapping or 'address' not in mapping:
        print(f"WARNING: Could not find required columns. Headers: {headers}")
        return []

    permits = []

    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        row = [cell.value for cell in sheet[row_idx]]

        if not any(row):
            continue

        permit_id = row[mapping['permit_id']] if 'permit_id' in mapping else None
        address = row[mapping['address']] if 'address' in mapping else None

        if not permit_id or not address:
            continue

        issued_date = None
        if 'date' in mapping and row[mapping['date']]:
            date_val = row[mapping['date']]
            if isinstance(date_val, datetime):
                issued_date = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, str):
                for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y']:
                    try:
                        issued_date = datetime.strptime(date_val.strip(), fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue

        permit = {
            'permit_id': str(permit_id).strip(),
            'address': str(address).strip(),
            'city': 'Ellis County (Unincorporated)',
            'date': issued_date,
            'type': str(row[mapping['type']]).strip() if 'type' in mapping and row[mapping['type']] else None,
            'value': row[mapping['value']] if 'value' in mapping else None,
            'owner_name': str(row[mapping['owner']]).strip() if 'owner' in mapping and row[mapping['owner']] else None,
            'contractor': str(row[mapping['contractor']]).strip() if 'contractor' in mapping and row[mapping['contractor']] else None,
        }
        permits.append(permit)

    return permits


def download_file(url: str, dest: Path) -> bool:
    """Download file from URL."""
    try:
        response = requests.get(url, timeout=60)
        response.raise_for_status()
        dest.write_bytes(response.content)
        return True
    except requests.RequestException as e:
        print(f"ERROR: Download failed: {e}")
        return False


def main() -> int:
    parser = argparse.ArgumentParser(description='Scrape Ellis County permits from Excel files')
    parser.add_argument('--url', help='URL of Excel file to download')
    parser.add_argument('--file', '-f', help='Local Excel file to parse')
    parser.add_argument('--discover', action='store_true', help='Show where to find Excel files')
    args = parser.parse_args()

    if args.discover or (not args.url and not args.file):
        print("Ellis County Building Permits:")
        print(f"  Portal: https://www.co.ellis.tx.us/index.aspx?nid=1074")
        print(f"  Reports: Look for 'Monthly Building Permit Data' Excel downloads")
        print("\nNOTE: This only covers UNINCORPORATED Ellis County.")
        print("Cities (Waxahachie, Ennis, Midlothian) have their own portals.")
        print("\nUsage:")
        print("  python3 scrapers/ellis_county_excel.py --file permits.xlsx")
        print("  python3 scrapers/ellis_county_excel.py --url https://...")
        return 0

    print('=' * 60)
    print('ELLIS COUNTY PERMIT SCRAPER (Excel)')
    print('=' * 60)
    print(f'Time: {datetime.now().isoformat()}\n')

    if args.file:
        input_file = Path(args.file)
        if not input_file.exists():
            print(f"ERROR: File not found: {input_file}")
            return 1
    elif args.url:
        input_file = OUTPUT_DIR / 'ellis_county_temp.xlsx'
        print(f'[1] Downloading from {args.url}...')
        if not download_file(args.url, input_file):
            return 1

    print(f'[2] Parsing {input_file.name}...')
    permits = parse_excel_permits(input_file)
    print(f'    Found {len(permits)} permits')

    if not permits:
        print("No permits found. Check file format.")
        return 1

    types = {}
    for p in permits:
        t = p['type'] or 'Unknown'
        types[t] = types.get(t, 0) + 1

    print('\n[3] Permits by type:')
    for ptype, count in sorted(types.items(), key=lambda x: -x[1])[:10]:
        print(f'    {ptype}: {count}')

    output_file = OUTPUT_DIR / "ellis_county_raw.json"

    output = {
        'source': 'ellis_county',
        'portal_type': 'Excel',
        'data_source': 'ellis_county_excel',
        'scraped_at': datetime.now().isoformat(),
        'actual_count': len(permits),
        'note': 'Unincorporated Ellis County only - not cities',
        'permits': permits
    }

    output_file.write_text(json.dumps(output, indent=2))

    print(f'\n' + '=' * 60)
    print('SUMMARY')
    print('=' * 60)
    print(f'Permits: {len(permits)}')
    print(f'Output: {output_file}')

    if permits:
        print('\nSAMPLE:')
        for p in permits[:3]:
            addr = p["address"][:40] if p["address"] else "N/A"
            print(f'  {p["date"]} | {p["permit_id"]} | {p["type"]} | {addr}')

    return 0


if __name__ == '__main__':
    main()
